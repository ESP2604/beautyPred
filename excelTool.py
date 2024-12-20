from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io, os
from openpyxl import Workbook ,load_workbook
import cv2
import numpy as np
import re
def convert_cv2_to_pil(cv2_image):
    """將OpenCV圖像轉換為PIL圖像"""
    # OpenCV存儲圖像的格式是BGR，轉換為RGB格式
    rgb_image = cv2.cvtColor(cv2_image, cv2.COLOR_BGR2RGB)
    # 將NumPy數組轉換為PIL圖像
    pil_image = PILImage.fromarray(rgb_image)
    return pil_image

def resize_and_compress_image_in_memory(img, scale_factor, quality=85):
    # 加载原始图像
    # img = PILImage.open(input_image_path)
    
    # 如果图像是RGBA模式，则转换为RGB
    if img.mode == 'RGBA':
        img = img.convert('RGB')
    
    # 计算新的尺寸
    new_width = int(img.width * scale_factor)
    new_height = int(img.height * scale_factor)
    
    # 调整图像尺寸
    img_resized = img.resize((new_width, new_height), PILImage.ANTIALIAS)
    
    # 保存调整后的图像到内存
    img_byte_arr = io.BytesIO()
    img_resized.save(img_byte_arr, format='JPEG', quality=quality)
    
    # 将内存中的图像数据定位回起始位置
    img_byte_arr.seek(0)
    
    return img_byte_arr
    
def insertImage(ws, img ,anchor ,row, scale_factor =1,quality=20):
        # 设置单元格的宽度和高度
    desired_cell_width =  50 # 根据需要调整
    desired_cell_height = 150  # 根据需要调整
    ws.column_dimensions[anchor[0]].width = desired_cell_width
    # for row in range(1, 5):  # 假设您要调整前四行的高度
    ws.row_dimensions[row].height = desired_cell_height

    # 加载并添加图片
    # img_path = "path_to_your_image.png"  # 替换为您的图片路径
    img = convert_cv2_to_pil(img)
    img_byte_arr = resize_and_compress_image_in_memory(img, 
        scale_factor=scale_factor, quality=quality)  # 根据需要调整scale_factor和quality
    img = Image(img_byte_arr)

    # 调整图片大小以适应单元格
    cell_ratio = desired_cell_height / desired_cell_width
    img_ratio = img.height / img.width
    if img_ratio > cell_ratio:
        # 图片更高，以高度为准进行缩放
        scale = (desired_cell_height*2) / img.height
    else:
        # 图片更宽，以宽度为准进行缩放
        scale = (desired_cell_width*3) / img.width

    img.width *= scale
    img.height *= scale

    # 将图片放置到特定的位置（例如，A1单元格）
    img.anchor = anchor

    # 将图片添加到工作表
    ws.add_image(img)
    


class ExcelSaver:
    def __init__(self, filePath='output.xlsx', sheetName='Sheet'):
        self.filePath = filePath
        self.sheetName = sheetName
        self.wb = None
        self.ws = None
        self.load_or_create_workbook()
    
    def load_or_create_workbook(self):
        """加載或創建工作簿和工作表"""
        if os.path.exists(self.filePath):
            self.wb = load_workbook(self.filePath)
        else:
            self.wb = Workbook()
        
        if self.sheetName in self.wb.sheetnames:
            self.ws = self.wb[self.sheetName]
        else:
            self.ws = self.wb.create_sheet(self.sheetName)

    def reTrySave(self, count):
        """重試保存工作簿"""
        reTry = 0
        while reTry < count:
            try:
                filename = os.path.basename(self.filePath).replace(".xlsx", "")
                parentdir = os.path.dirname(self.filePath)
                self.wb.save(f'{os.path.join(parentdir, filename)}({reTry}).xlsx')
                break  # 如果保存成功，退出循环
            except PermissionError:
                reTry += 1  # 增加重试次数
            except Exception as e:
                print(f"保存失败：{e}")
                return  # 如果遇到其他异常，停止重试并退出

        if reTry >= count:
            print("重试次数已达上限，保存失败。")

    def writeExcel(self , data,i,):
        for item in data:
            value = item['value']
            column = item['column']
            row = item['row']
            # 如果值是URL，插入超链接
            if isinstance(value, str) and self.is_url(value):
                # 插入超链接
                self.ws[f'{column}{row}'] = value
                self.ws[f'{column}{row}'].hyperlink = value
                self.ws[f'{column}{row}'].style = 'Hyperlink'  # 设置超链接样式
            # 如果值是NumPy数组或PIL图像，插入图片
            elif isinstance(value, np.ndarray) or isinstance(value, PILImage.Image):
                insertImage(self.ws, value, f'{column}{row}', row, scale_factor=1, quality=60)
            else:
                # 否则，插入普通的文本数据
                self.ws[f'{column}{row}'] = value

    def is_url(self, string):
        """检测字符串是否是URL"""
        url_regex = re.compile(
            r'^(?:http|ftp)s?://' # 匹配http:// 或 https://
            r'(?:(?:[A-Z0-9](?:[A-Z0-9-]*[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' # 匹配域名
            r'localhost|' # 本地主机名
            r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}|' # 匹配IP地址
            r'\[?[A-F0-9]*:[A-F0-9:]+\]?)' # 匹配IPv6地址
            r'(?::\d+)?' # 匹配端口号
            r'(?:/?|[/?]\S+)$', re.IGNORECASE)
        
        return re.match(url_regex, string) is not None