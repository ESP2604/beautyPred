import cv2
import numpy as np
from PIL import Image
import torch
from efficientnet_pytorch import EfficientNet
import torch.nn as nn
import torchvision.transforms as transforms
import glob
import os
from openpyxl import Workbook ,load_workbook
import excelTool
# from openpyxl.drawing.image import Image as OpenpyxlImage
import argparse
import MyImageTool
# 定义要检查的目录路径
directory_path = 'tmp'

# 检查目录是否存在
if not os.path.exists(directory_path):
    try:
        # 如果目录不存在，创建它
        os.makedirs(directory_path)
        print(f"目录 '{directory_path}' 已创建。")
    except OSError as e:
        print(f"创建目录 '{directory_path}' 时出错: {e}")
else:
    print(f"目录 '{directory_path}' 已存在。")

# filePath 檔案路徑
# count 最重試次數
def reTrySave(filePath, count):
    reTry = 0
    while reTry < count:
        try:
            filename = os.path.basename(filePath).replace(".xlsx", "")
            parentdir = os.path.dirname(filePath)
            # 保存工作簿
            wb.save(f'{os.path.join(parentdir, filename)}({reTry}).xlsx')
            break  # 如果保存成功，退出循环
        except PermissionError:
            reTry += 1  # 增加重试次数
        except Exception as e:
            print(f"保存失败：{e}")
            return  # 如果遇到其他异常，停止重试并退出

    if reTry >= count:
        print("重试次数已达上限，保存失败。")

def writeExcel(data,i):
    # 写入分数和图像名称
    ws[f'A{i}'] = data.score
    ws[f'B{i}'] = data.image_name
    ws[f'F{i}'] = data.padding_score
    # 插入图像
    if os.path.exists(data.image_path):
        excelTool.insertImage(ws, data.image_path,f'C{data.row}', data.row ,scale_factor =1,quality=20)
    # 插入头像
    if os.path.exists(data.avatar_path):
        excelTool.insertImage(ws, data.avatar_path,f'D{data.row}' , data.row,scale_factor =1,quality=20)
    # 插入padding過後的頭向
    if os.path.exists(data.padding_avatar_path):
        excelTool.insertImage(ws, data.padding_avatar_path,f'E{data.row}' , data.row,scale_factor =1,quality=20)


class AvatarImage:
    def __init__(self, image,coords) -> None:
        self.coords = coords
        self.originalImage = image.copy()
        self.processedImage = image.copy()
        self.infoImage = image.copy()
        self.score = 0
        # 裁減大頭貼
        new_x, new_y, new_w, new_h = self.expand_coords()
        self.faceImage = self.originalImage[ new_y:new_y  + new_h, new_x:new_x + new_w],
        # "原始圖片"翻譯成英文是"Original Image"。
        # "加工圖片"翻譯成英文是"Processed Image"。
    def printScore(self, score):
        height, width = self.infoImage.shape[:2]
        # Put score
        cv2.putText(self.infoImage, 'Score:{:.4f}'.format(score), ( 0, height - 30), cv2.FONT_HERSHEY_SIMPLEX, 1,
                    (255, 0, 0) )
        return self

    def paddingReSetRect( self, start_x ,start_y):
        new_x, new_y, new_w, new_h = self.expand_coords()
        if max(0,new_x - start_x) == 0 : new_x = 0
        if max(0,new_y - start_y) == 0 : new_y = 0
        # 绘制矩形框
        cv2.rectangle(self.infoImage, (new_x, new_y), (new_x + new_w, new_y + new_h), (0, 255, 0), 2)
        return self

    def facerect(self):
        """
            標記臉部框框
        """
        # 假设 coords 是一个包含四个元素的列表或元组，格式为 [x, y, width, height]
        new_x, new_y, new_w, new_h = self.expand_coords()
        # 绘制矩形框
        cv2.rectangle(self.infoImage, (new_x, new_y), (new_x + new_w, new_y + new_h), (0, 255, 0), 2)
        return self

    def markFeatures(self):
        '''
            標記五官特徵
        '''
        # 標記五官
        cv2.circle(self.infoImage, (self.coords[4], self.coords[5]), 2, (255, 0, 0), 2)
        cv2.circle(self.infoImage, (self.coords[6], self.coords[7]), 2, (0, 0, 255), 2)
        cv2.circle(self.infoImage, (self.coords[8], self.coords[9]), 2, (0, 255, 0), 2)
        cv2.circle(self.infoImage, (self.coords[10], self.coords[11]), 2, (255, 0, 255), 2)
        cv2.circle(self.infoImage, (self.coords[12], self.coords[13]), 2, (0, 255, 255), 2)
        return self
    def expand_coords(self):
        """
        根据给定的扩展比例放大坐标
        :param coords: 包含四个元素的列表或元组，格式为 [x, y, width, height]
        :return: 放大后的新坐标 [new_x, new_y, new_w, new_h]
        """
            # 假设 coords 是一个包含四个元素的列表或元组，格式为 [x, y, width, height]
        x = self.coords[0]
        y = self.coords[1]
        w = self.coords[2] 
        h = self.coords[3] 
        width= self.infoImage.shape[0]
        height = self.infoImage.shape[1]
        # 計算可以用於裁剪的最大正方形大小
        square_size = min([width - x, x, y, height - y])

        # 確保square_size為正，並且在圖像邊界內
        square_size = max(min(square_size, w, h), 0)
        new_x= max(x - square_size, 0)
        new_y= max(y - square_size, 0)
        new_w=  (x + w + square_size) - max(x - square_size, 0) 
        new_h=  (y + h + square_size) - max(y - square_size, 0) 
        return int(new_x), int(new_y), int(new_w), int(new_h)

class CropFace:
    def __init__(self):
        self.yunet = CropFace.createFaceDetectorYN()
    
    def createFaceDetectorYN(modelStr = '.\\face_detection_yunet_2022mar.onnx'):
        score_threshold = 0.85
        nms_threshold = 0.35
        backend = cv2.dnn.DNN_BACKEND_DEFAULT
        target = cv2.dnn.DNN_TARGET_CPU
        # Instantiate yunet
        yunet = cv2.FaceDetectorYN.create(
            model=modelStr,
            config='',
            input_size=(320, 320),
            score_threshold=score_threshold,
            nms_threshold=nms_threshold,
            top_k=5000,
            backend_id=backend,
            target_id=target
        )
        return yunet

    def detect(self,image):
        self.yunet.setInputSize((image.shape[1], image.shape[0]))
        _, faces = self.yunet.detect(image)  # faces: None, or nx15 np.array
        list = []
        for idx, face in enumerate(faces):
            coords = face[:-1].astype(np.int32)
            list.append(AvatarImage(image, coords))
        return list

class FaceRating:
    """
        # 读取图像
        image = cv2.imread(img_file)
        # 圖片 加上padding
        # 判斷圖片
        actorImages = cropFace.detect(image)
        # 要標記的東西
        for item in actorImages:
            # 捕捉到的臉部標記
            item.facerect()
        if len(actorImages) > 0: 
            original_images = [actor_image.originalImage for actor_image in actorImages]
            scores = faceRating.pre(original_images)
            print(scores)
            if not isinstance(scores, (list, np.ndarray)):
                scores = [scores]

            # 现在可以安全地迭代scores了
            for idx, score in enumerate(scores):
                actorImages[idx].score = score
    """
    def __init__(self):
        # 1. 检查GPU是否可用
        if torch.cuda.is_available():
            self.device = torch.device("cuda")
        else:
            self.device = torch.device("cpu")
        self.model = self.loadModel()
        self.transform = transforms.Compose([
        transforms.Resize(256),
            transforms.CenterCrop(224),
            transforms.ToTensor(),
            transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
        ])
    def loadModel(self):
        # 加载预训练的 EfficientNet 模型
        model = EfficientNet.from_pretrained('efficientnet-b0')  
        # 修改最后一层以适应回归任务
        model._fc = nn.Linear(model._fc.in_features, 1)
        model.load_state_dict(torch.load('models/efficient2_model_epoch_15.pth'))
        model.to(self.device)
        model.eval()
        return model

    # def pre(self ,img):
    #     # 将 NumPy 数组转换为 PIL 图像
    #     image_pil = Image.fromarray(img)
    #     image_tensor =self.transform(image_pil).to(self.device)
    #     image_tensor =image_tensor.unsqueeze(0)
    #     with torch.no_grad():
    #         prediction = self.model(image_tensor).cpu().item()
    #         return prediction
    def pre(self, imgs):
        # 统一处理单张图片和图片批量
        if isinstance(imgs, np.ndarray):  # 检查是否为单张图片
            imgs = [imgs]  # 将单张图片转换为列表
        
        images_pil = [Image.fromarray(img) for img in imgs]  # 将每个 NumPy 数组转换为 PIL 图像
        images_tensor = torch.stack([self.transform(image) for image in images_pil]).to(self.device)
        
        with torch.no_grad():
            predictions = self.model(images_tensor).cpu().numpy().flatten()
            
        if len(predictions) == 1:
            return predictions[0]  # 如果是单张图片，返回单个预测值
        return predictions  # 如果是图片批量，返回预测值数组

    def preBetch(self, imgs):
        # 适用于批量预测的方法
        images_pil = [Image.fromarray(img) for img in imgs]  # 将每个 NumPy 数组转换为 PIL 图像
        images_tensor = torch.stack([self.transform(image) for image in images_pil]).to(self.device)
        
        with torch.no_grad():
            predictions = self.model(images_tensor).cpu().numpy().flatten()
            return predictions

  
  
    
# 創建 ArgumentParser 對象
parser = argparse.ArgumentParser(description='人臉打分輸出excel')

# 添加參數，並為部分參數設置預設值
parser.add_argument('--excel', nargs='?', default='output1.xlsx', help='輸出excel路徑，如果有則寫入excel的其他分頁，預設為output.xlsx')
parser.add_argument('--sheetname', nargs='?', default='Sheet', help='excel 分頁名稱，預設為sheet')
parser.add_argument('--source', nargs='?' , default='TestImage', help='圖片資料夾，預設TestImage')
parser.add_argument('--score', nargs='?', type=int, default=1, help='分數[1~5]，大於輸入的值才添加進excel，預設為1')
parser.add_argument('--limit', nargs='?', type=int, default=200, help='最大處理圖片數量 預設為200')
# 解析參數
args = parser.parse_args()
limit = args.limit
# 设定文件夹路径
folder_path = args.source

# 创建一个工作簿
filename = args.excel
sheet_name = args.sheetname
wb = None
ws = None
if os.path.exists(filename):
    wb = load_workbook(filename)
else:
    wb = Workbook()
# 检查工作簿中是否存在名为 "sheetname" 的工作表
if sheet_name in wb.sheetnames:
    # print("工作表已经存在")
    ws = wb[sheet_name]
else:
    # print("工作表不存在")
    ws = wb.create_sheet(sheet_name)
i = 1
imagePaths =  glob.glob(folder_path + '/*.jpg') + glob.glob(folder_path + '/*.png') +  glob.glob(folder_path + '/*.jpeg')
limit = min(limit, len(imagePaths))
cropFace = CropFace()
faceRating = FaceRating()
# 遍历文件夹中的所有图片
for img_file in imagePaths:
     # 读取图像
    image = cv2.imread(img_file)
    # 圖片 加上padding
    # 判斷圖片
    actorImages = cropFace.detect(image)
    # 要標記的東西
    for item in actorImages:
        # 捕捉到的臉部標記
        item.facerect()
    if len(actorImages) > 0: 
        original_images = [actor_image.originalImage for actor_image in actorImages]
        scores = faceRating.pre(original_images)
        print(scores)
        if not isinstance(scores, (list, np.ndarray)):
            scores = [scores]

        # 现在可以安全地迭代scores了
        for idx, score in enumerate(scores):
            actorImages[idx].score = score
    
    padding_image, start_y , new_height, start_x ,  new_width = MyImageTool.create_centered_resized_image(image)
    

    paddingActorImages = cropFace.detect(padding_image)
    for item in paddingActorImages:
        item.facerect()

    if len(paddingActorImages) > 0: 
        original_images = [actor_image.originalImage for actor_image in paddingActorImages]
        scores = faceRating.pre(original_images)
        # 确保scores是一个列表
        if not isinstance(scores, (list, np.ndarray)):
            scores = [scores]

        # 现在可以安全地迭代scores了
        for idx, score in enumerate(scores):
            paddingActorImages[idx].score = score


    if(actorImages is None  or paddingActorImages == 0 ): continue 
    # padding_vis_image = image.copy()
    # 通常只有一張臉
    # paddingReSetRect(padding_vis_image,faceImgList[0]['coords'], start_x, start_y)

    # 将结果图像保存
    save_path = os.path.join('tmp/', os.path.basename(img_file))
    # save_path=""
    cv2.imwrite(save_path, actorImages[0].infoImage)
    # 将结果图像保存
    padding_save_path = os.path.join('tmp/', f'padding_{os.path.basename(img_file)}')

    cv2.imwrite(padding_save_path, paddingActorImages[0].infoImage)

    # score = padding_score
    data = excelTool.MyExcelData()
    data.score = actorImages[0].score
    data.padding_score = paddingActorImages[0].score
    data.image_path = img_file
    data.image_name = os.path.basename(img_file)
    data.avatar_path = save_path
    data.padding_avatar_path = padding_save_path
    data.row = i
    writeExcel(data,i)
    i+=1
    if(i >limit): break
    
    print(f'{i}/{limit}分数[1~5]: original:{data.score} padding:{data.padding_score}')
    # cv2.imshow('xx', vis_image)
    # cv2.waitKey(0)
reTrySave("output.xlsx",  10)